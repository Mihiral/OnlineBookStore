from django.shortcuts import render,redirect
from django.template import RequestContext
from django.contrib import messages
import pymysql
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
import datetime
import os
from datetime import date
import base64
import os
import logging
from openpyxl import Workbook


global uname, query
shopList = []

def RatingsAction(request):
    if request.method == 'POST':
        global uname
        bid = request.POST.get('bid', False)
        ratings = request.POST.get('ratings', False)
        feedback = request.POST.get('feedback', False)
        today = date.today()
        db_connection = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        db_cursor = db_connection.cursor()
        student_sql_query = "INSERT INTO feedback(username,book_name,ratings,feedback,feedback_date) VALUES('"+uname+"','"+bid+"','"+ratings+"','"+feedback+"','"+str(today)+"')"
        db_cursor.execute(student_sql_query)
        db_connection.commit()
        print(db_cursor.rowcount, "Record Inserted")
        if db_cursor.rowcount == 1:
            status = 'Your feedback accepted'
        context= {'data':status}
        return render(request, 'UserScreen.html', context)

def Ratings(request):
    if request.method == 'GET':
        output = ''
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select book_name from addbook")
            rows = cur.fetchall()
            for row in rows:
                output += '<option value="'+row[0]+'">'+row[0]+'</option>'
        context= {'data1': output}
        return render(request, 'Ratings.html', context)

def CancelOrder(request):
    if request.method == 'GET':
        orderid = request.GET.get('t1', False)
        db_connection = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        db_cursor = db_connection.cursor()
        student_sql_query = "update orders_new set order_status='Cancelled' where order_id='"+orderid+"'"
        db_cursor.execute(student_sql_query)
        db_connection.commit()
        status = 'Ordered Cancelled Successfully'
        context= {'data':status}
        return render(request, 'UserScreen.html', context)

def ReviewOrders(request):
    if request.method == 'GET':
        global uname
        output = '<table border=1><tr><th><font size="" color=white>Order ID</font></th>'
        output+='<td><font size="" color="white">Username</td>'
        output+='<td><font size="" color="white">Ordered Date</td>'
        output+='<td><font size="" color="white">Purchased Book ID</td>'
        output+='<td><font size="" color="white">Total Amount</td>'
        output+='<td><font size="" color="white">Order Status</td>'
        output+='<td><font size="" color="white">Cancel Order</td></tr>'
        rank = []
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * from orders_new where username='"+uname+"'")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                output+='<td><font size="" color="white">'+str(row[4])+'</td>'
                output+='<td><font size="" color="white">'+str(row[5])+'</td>'
                if row[5] == 'Pending':
                    output+='<td><a href=\'CancelOrder?t1='+str(row[0])+'\'><font size=3 color=white>Click Here to Cancel Order</font></a></td></tr>'
                else:
                    output+='<td><font size="" color="white">'+str(row[5])+'</td></tr>'
        output += "</table><br/><br/><br/>"
        context= {'data': output}
        return render(request, 'Generic.html', context)

def ViewShippingDetails(request):
    if request.method == 'GET':
        global uname
        output = '<table><tr><th>Username</th><th>Shipping Address</th></tr>'
        rank = []
        con = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='root',
                              database='onlinebookstore', charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("SELECT address FROM newuser WHERE username=%s", (uname,))
            rows = cur.fetchall()
            for row in rows:
                output += '<tr><td>{}</td><td>{}</td></tr>'.format(uname, row[0])
        output += "</table><br/><br/>"
        context = {'data': output}
        return render(request, 'ShippingDetails.html', context) 

def SearchBook(request):
    if request.method == 'GET':
        return render(request, 'SearchBook.html', {})

def getCost(book_id):
    cost = 0
    con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
    with con:
        cur = con.cursor()
        cur.execute("select cost from addbook where book_id='"+book_id+"'")
        rows = cur.fetchall()
        for row in rows:
            cost = str(row[0])
    return float(cost)

def PaymentAction(request):
    if request.method == 'POST':
        # Get the user's shopping cart from the session
        shopList = request.session.get('shopList', [])
        global uname
        #uname = request.session.get('uname', '')

        details = ""
        total = 0
        for book_id in shopList:
            total += getCost(book_id)
            details += book_id + ","

        # Clear the shopping cart in the session
        request.session['shopList'] = []

        details = details[0:len(details)-1]
   
        order_id = 0
        con = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='root', database='onlinebookstore', charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select max(order_id) from orders_new")
            rows = cur.fetchall()
            for row in rows:
                order_id = row[0]
        if order_id is not None:
            order_id = order_id + 1
        else:
            order_id = 1
        today = date.today()
        db_connection = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='root', database='onlinebookstore', charset='utf8')
        db_cursor = db_connection.cursor()
        student_sql_query = "INSERT INTO orders_new(order_id,username,order_date,product_id,total_amount,order_status) VALUES('"+str(order_id)+"','"+uname+"','"+str(today)+"','"+details+"','"+str(total)+"','Pending')"
        db_cursor.execute(student_sql_query)
        db_connection.commit()
        print(db_cursor.rowcount, "Record Inserted")
        if db_cursor.rowcount == 1:
            status = 'Order completed<br/>Total Amount to be paid after delivery : '+str(total)
        context = {'data': status}
        return render(request, 'Generic.html', context)

def Payment(request):
    if request.method == 'GET':
        # Get the user's shopping cart from the session
        shopList = request.session.get('shopList', [])

        output = '<tr><td><b>Username</b></td><td><input type="text" name="t3" style="font-family: Comic Sans MS" size="30" value="'+uname+'"/></td></tr>'
        output +='<tr><td><b>Expiry&nbsp;Date</b></td><td><select name="t4">'
        for i in range(2023, 2050):
            output += '<option value="'+str(i)+'">'+str(i)+'</option>'
        output += '</select>&nbsp;<select name="t5">'
        for i in range(1, 13):
            output += '<option value="'+str(i)+'">'+str(i)+'</option>'
        output += "</select></td></tr>"

        # Pass the shopping cart data to the template
        context = {'data1': output, 'shopList': shopList}
        return render(request, 'Payment.html', context)    

def BackToCart(request):
    if request.method == 'GET':
        global shopList, query
        output = '<table border=1><tr><th><font size="" color=white>Book ID</font></th>'
        output+='<td><font size="" color="white">Book Name</td>'
        output+='<td><font size="" color="white">Category</td>'
        output+='<td><font size="" color="white">Author Name</td>'
        output+='<td><font size="" color="white">Subject</td>'
        output+='<td><font size="" color="white">Cost</td>'
        output+='<td><font size="" color="white">Image</td>'
        output+='<td><font size="" color="white">Add to Cart</td>'
        output+='<td><font size="" color="white">View Cart</td></tr>'
        rank = []
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        
        if 'query' not in locals() and 'query' not in globals():
            query = ""
        with con:
            cur = con.cursor()
            
            cur.execute("select * from addbook where book_name='"+query+"' or category='"+query+"' or author_names='"+query+"'")
            rows = cur.fetchall()
            if not rows:
                context = {'message': 'Find your favourite book of your choice!'}
                return render(request, 'SearchBook.html', context)
                #print("No search results found. Go to search to find your favourite book of your choice!")
            else:
              for row in rows:
                  output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                  output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                  output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                  output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                  output+='<td><font size="" color="white">'+str(row[4])+'</td>'
                  output+='<td><font size="" color="white">'+str(row[5])+'</td>'
                  output+='<td><img src="static/books/'+row[6]+'" width="100" height="100" alt="" /></td>'
                  output+='<td><a href=\'AddCart?t1='+str(row[0])+'\'><font size=3 color=white>Add to Cart</font></a></td>'
                  output+='<td><a href=\'ViewCart\'><font size=3 color=white>View Cart</font></a></td></tr>'
        output += "</table><br/><br/>"
        if len(shopList) > 0:
            output += '<a href=\'Payment\'><font size=3 color=white>Confirm Order</font></a>'
        context= {'data': output}
        return render(request, 'Generic.html', context)    

def ViewCart(request):
    if request.method == 'GET':
        shopList = request.session.get('shopList', [])
        query = request.POST.get('t1', False)
        output = '<table border=1><tr><th><font size="" color=white>Serial No</font></th>'
        output+='<td><font size="" color="white">Book ID</td>'
        output+='<td><font size="" color="white">Book Name</td>'
        output+='<td><font size="" color="white">Category</td>'
        output+='<td><font size="" color="white">Author Name</td>'
        output+='<td><font size="" color="white">Subject</td>'
        output+='<td><font size="" color="white">Cost</td>'
        output+='<td><font size="" color="white">Image</td></tr>'

        if not shopList:
            context = {'data': "Cart is empty! Please search for the books you want"}
            return render(request, 'SearchBook.html', context)
        
        for i, book_id in enumerate(shopList):
            con = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='root', database='onlinebookstore', charset='utf8')
            with con:
                cur = con.cursor()
                cur.execute("select * from addbook where book_id=%s", (book_id,))
                rows = cur.fetchall()
            
                for row in rows:
                    output+='<tr><td><font size="" color="white">'+str(i+1)+'</td>'
                    output+='<td><font size="" color="white">'+str(row[0])+'</td>'
                    output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                    output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                    output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                    output+='<td><font size="" color="white">'+str(row[4])+'</td>'
                    output+='<td><font size="" color="white">'+str(row[5])+'</td>'
                    output+='<td><img src="static/books/'+row[6]+'" width="100" height="100" alt="" /></td></tr>'
        output += "</table><br/><br/>"
        if len(shopList) > 0:
            output += '<a href=\'Payment\'><font size=3 color=white>Confirm Order</font></a>'
        context= {'data': output}
        return render(request, 'Generic.html', context)    

def AddCart(request):
    if request.method == 'GET':
        book_id = request.GET.get('t1', '')
        shopList = request.session.get('shopList', [])
        shopList.append(book_id)
        request.session['shopList'] = shopList
        
        output =  '<h2><p> Review Item Details </p></h2>' 
        output += '<table border=1><tr><th><font size="" color=white>Book ID</font></th>'
        output += '<td><font size="" color="white">Book Name</td>'
        output += '<td><font size="" color="white">Category</td>'
        output += '<td><font size="" color="white">Author Name</td>'
        output += '<td><font size="" color="white">Subject</td>'
        output += '<td><font size="" color="white">Cost</td>'
        output += '<td><font size="" color="white">Image</td>'

        con = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='root', database='onlinebookstore', charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * from addbook where book_id=%s", (book_id,))
            rows = cur.fetchall()
            for row in rows:
                output += f'<tr><td><font size="" color="white">{str(row[0])}</td>'
                output += f'<td><font size="" color="white">{str(row[1])}</td>'
                output += f'<td><font size="" color="white">{str(row[2])}</td>'
                output += f'<td><font size="" color="white">{str(row[3])}</td>'
                output += f'<td><font size="" color="white">{str(row[4])}</td>'
                output += f'<td><font size="" color="white">{str(row[5])}</td>'
                output += f'<td><img src="static/books/{row[6]}" width="100" height="100" alt="" /></td>'

        output += "</table><br/><br/>"
        if shopList:
            output += '<a href=\'ViewCart\'><font size=3 color=white>Go to Cart</font></a>&nbsp;&nbsp;&nbsp;&nbsp;'
            output += '<a href="Payment"><font size=3 color=white>Confirm Order</font></a>'

        context = {'data': output}
        return render(request, 'ViewCart.html', context)       
    

def SearchBookAction(request):
    if request.method == 'POST':
        query = request.POST.get('t1', '')
        output = ''

        con = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='root', database='onlinebookstore', charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * from addbook where book_name LIKE %s OR category LIKE %s OR author_names LIKE %s",
                        ('%' + query + '%', '%' + query + '%', '%' + query + '%'))
            rows = cur.fetchall()

            if not rows:
                context = {'data': "No results found! Please search for another title or keyword!"}
                return render(request, 'SearchBook.html', context)

            for row in rows:
                output += f'<p><img src="static/books/{row[6]}" width="200" height="200" alt="" /></p></font>'
                output += f'<p><font size="" color="white">{str(row[1])}</p>'
                output += f'<p><font size="" color="white">{str(row[3])}</p>'
                output += f'<p><font size="" color="white">{str(row[5])}</p>'
                output += f'<p><a href="AddCart?t1={str(row[0])}"><font size=3 color=white>Add to Cart</font></a></p>'
                output += '<p><a href="ViewCart"><font size=3 color=white>View Cart</font></a></p>'

        output += "<br/><br/>"
        context = {'data': output}
        return render(request, 'Generic.html', context) 

        #output = '<table border=1><tr><th><font size="" color=white>Book ID</font></th>'
        #output+='<td><font size="" color="white">Book Name</td>'
        #output+='<td><font size="" color="white">Category</td>'
        #output+='<td><font size="" color="white">Author Name</td>'
        #output+='<td><font size="" color="white">Subject</td>'
       # output+='<td><font size="" color="white">Cost</td>'
        #output+='<td><font size="" color="white">Image</td>'
        #output+='<td><font size="" color="white">Add to Cart</td>'
        #output+='<td><font size="" color="white">View Cart</td></tr>'
        #rank = []

        
        #con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        #with con:
        #    cur = con.cursor()
        #    cur.execute("select * from addbook where book_name LIKE %s OR category LIKE %s OR author_names LIKE %s", ('%' + query + '%', '%' + query + '%', '%' + query + '%'))
        #               # = '"+query+"' or category='"+query+"' or author_names='"+query+"'")
        #    rows = cur.fetchall()
       #     for row in rows:
        #        output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
        #        output+='<td><font size="" color="white">'+str(row[1])+'</td>'
        #        output+='<td><font size="" color="white">'+str(row[2])+'</td>'
        #        output+='<td><font size="" color="white">'+str(row[3])+'</td>'
        #        output+='<td><font size="" color="white">'+str(row[4])+'</td>'
        #        output+='<td><font size="" color="white">'+str(row[5])+'</td>'
        #        output+='<td><img src="static/books/'+row[6]+'" width="100" height="100" alt="" /></td>'
        #        output+='<td><a href=\'AddCart?t1='+str(row[0])+'\'><font size=3 color=white>Add to Cart</font></a></td>'
        #        output+='<td><a href=\'ViewCart\'><font size=3 color=white>View Cart</font></a></td></tr>'
       # output += "</table><br/><br/>"
       
        #if len(shopList) > 0:
        #    output += '<a href=\'Payment\'><font size=3 color=white></font></a>'
        #context= {'data': output}
       # return render(request, 'Generic.html', context)
    
def get_all_categories(request):
    con = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='root',
                          database='onlinebookstore', charset='utf8')
    categories = []

    with con:
        cur = con.cursor()
        cur.execute("SELECT DISTINCT category FROM addbook")
        rows = cur.fetchall()

        for row in rows:
            categories.append(row[0])

    return categories

def Login(request):
    if request.method == 'GET':
        return render(request, 'Login.html', {})

def index(request):
    if request.method == 'GET':
        return render(request, 'index.html', {})

def Register(request):
    if request.method == 'GET':
        return render(request, 'Register.html', {})

def Admin(request):
    if request.method == 'GET':
        return render(request, 'Admin.html', {})
    
def AdminHome(request):
    if request.method == 'GET':
        return render(request, 'AdminHome.html', {})
def UserHome(request):
    if request.method == 'GET':
        return render(request, 'UserScreen.html', {})

def RegisterAction(request):
    if request.method == 'POST':
        username = request.POST.get('t1', False)
        password = request.POST.get('t2', False)
        contact = request.POST.get('t3', False)
        address = request.POST.get('t4', False)
        email = request.POST.get('t5', False)
        
        status = 'none'
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select username from newuser where username = '"+username+"'")
            rows = cur.fetchall()
            for row in rows:
                if row[0] == email:
                    status = 'Given Username already exists'
                    break
        if status == 'none':
            db_connection = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
            db_cursor = db_connection.cursor()
            student_sql_query = "INSERT INTO newuser(username,password,contact_no,address,email) VALUES('"+username+"','"+password+"','"+contact+"','"+address+"','"+email+"')"
            db_cursor.execute(student_sql_query)
            db_connection.commit()
            print(db_cursor.rowcount, "Record Inserted")
            if db_cursor.rowcount == 1:
                status = 'Signup Process Completed'
        context= {'data':status}
        return render(request, 'Register.html', context)

def UserLoginAction(request):
    if request.method == 'POST':
        global uname
        option = 0
        username = request.POST.get('t1', False)
        password = request.POST.get('t2', False)
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * FROM newuser")
            rows = cur.fetchall()
            for row in rows:
                if row[0] == username and row[1] == password:
                    uname = username
                    option = 1
                    break
        if option == 1:
            context= {'data':'welcome '+username}
            return render(request, 'UserScreen.html', context)
        else:
            context= {'data':'Invalid login details'}
            return render(request, 'Login.html', context)

def AdminLoginAction(request):
    if request.method == 'POST':
        global uname
        option = 0
        username = request.POST.get('t1', False)
        password = request.POST.get('t2', False)
        if username == "admin" and password == "admin":
            context= {'data':'welcome '+username}
            return render(request, 'AdminHome.html', context)
        else:
            context= {'data':'Invalid login details'}
            return render(request, 'Admin.html', context)        

def AddBook(request):
    if request.method == 'GET':
        return render(request, 'AddBook.html', {})
    
def AddBookAction(request):
    if request.method == 'POST':
        global uname
        myfile = request.FILES['t6'].read()
        fname = request.FILES['t6'].name
        book = request.POST.get('t1', False)
        category = request.POST.get('t2', False)
        author = request.POST.get('t3', False)
        subject = request.POST.get('t4', False)
        cost = request.POST.get('t5', False)        
        with open("BookStoreApp/static/books/"+fname, "wb") as file:
            file.write(myfile)
        file.close()
        file_id = 0
        con = pymysql.connect(host='127.0.0.1', port = 3306, user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select max(book_id) from addbook")
            rows = cur.fetchall()
            for row in rows:
                file_id = row[0]
        if file_id is not None:
            file_id = int(file_id) + 1
        else:
            file_id = 1

        while True:
            with con:
                cur = con.cursor()
                cur.execute("SELECT 1 FROM addbook WHERE book_id = %s", (file_id,))
                if not cur.fetchone():
                    break  # Unique ID found
            file_id += 1
        db_connection = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        db_cursor = db_connection.cursor()
        student_sql_query = "INSERT INTO addbook(book_id,book_name,category,author_names,subject,cost,image) VALUES('"+str(file_id)+"','"+book+"','"+category+"','"+author+"','"+subject+"','"+cost+"','"+fname+"')"
        db_cursor.execute(student_sql_query)
        db_connection.commit()
        status = "Error in file upload"
        if db_cursor.rowcount == 1:
            status = "Book successfully uploaded to server"
        context= {'data':status+"<br/><br/><br/><br/><br/>"}
        return render(request, 'AddBook.html', context)

def DeleteAction(request):
    if request.method == 'GET':
        filename = request.GET.get('t1', False)
        db_connection = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        db_cursor = db_connection.cursor()
        student_sql_query = "delete from addbook where book_id = '"+filename+"'"
        db_cursor.execute(student_sql_query)
        db_connection.commit()
        status = "Error in file deleting"
        if db_cursor.rowcount == 1:
            status = "Book successfully deleted from Database"
        context= {'data':status+"<br/><br/><br/><br/><br/>"}
        return render(request, 'AdminScreen.html', context)

def Delete(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Book ID</font></th>'
        output+='<td><font size="" color="white">Book Name</td>'
        output+='<td><font size="" color="white">Category</td>'
        output+='<td><font size="" color="white">Author Name</td>'
        output+='<td><font size="" color="white">Subject</td>'
        output+='<td><font size="" color="white">Cost</td>'
        output+='<td><font size="" color="white">Image</td>'
        output+='<td><font size="" color="white">Delete Book</td></tr>'
        rank = []
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * FROM addbook")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                output+='<td><font size="" color="white">'+str(row[4])+'</td>'
                output+='<td><font size="" color="white">'+str(row[5])+'</td>'
                output+='<td><img src="static/books/'+row[6]+'" width="100" height="100" alt="" /></td>'
                output+='<td><a href=\'DeleteAction?t1='+str(row[0])+'\'><font size="" color=white>Click Here to Delete Book</font></a></td></tr>'
        output += "</table><br/><br/><br/>"
        context= {'data': output}
        return render(request, 'AdminScreen.html', context)
    
def UpdatePrice(request):
    if request.method == 'GET':
        output = '<tr><td><b>Book&nbsp;Name</b></td><td><select name="t1">'
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select book_name FROM addbook")
            rows = cur.fetchall()
            for row in rows:
                output += '<option value="'+row[0]+'">'+row[0]+'</option>'
        output += '</select></td></tr>'
        context= {'data': output}
        return render(request, 'UpdatePrice.html', context)


def UpdatePriceAction(request):
    if request.method == 'POST':
        book = request.POST.get('t1', False)
        price = request.POST.get('t2', False)
        db_connection = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        db_cursor = db_connection.cursor()
        student_sql_query = "update addbook set cost='"+price+"' where book_name='"+book+"'"
        db_cursor.execute(student_sql_query)
        db_connection.commit()
        status = 'Book price updated successfully'
        context= {'data':status}
        return render(request, 'AdminScreen.html', context)

def ViewPurchase(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Order ID</font></th>'
        output+='<td><font size="" color="white">Username</td>'
        output+='<td><font size="" color="white">Ordered Date</td>'
        output+='<td><font size="" color="white">Purchased Book ID</td>'
        output+='<td><font size="" color="white">Total Amount</td>'
        output+='<td><font size="" color="white">Order Status</td></tr>'
        rank = []
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * from orders_new")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                output+='<td><font size="" color="white">'+str(row[4])+'</td>'
                output+='<td><font size="" color="white">'+str(row[5])+'</td>'
                '</td></tr>'
        output += "</table><br/><br/><br/>"
        context= {'data': output}
        return render(request, 'AdminScreen.html', context)
    
def AvgRatings(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Book Name</font></th>'
        output+='<td><font size="" color="white">Rating</td>'
        rank = []
        rows = ""
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select book_name, avg(ratings) from feedback group by book_name")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                '</td></tr>'
        output += "</table><br/><br/><br/>"
        wb = Workbook()
        ws = wb.active

    # Adding column headers
        headers = ['book_name', 'Average Ratings']
        ws.append(headers)

    # Adding data rows
        for row in rows:
            ws.append(row)

    # Create the response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=exported_data.xlsx'

    # Save the workbook to the response
        wb.save(response)
        context= {'data': output, 'excel_link': response}
        return render(request, 'ratingsreport.html', context)  
    
def AvgRatingsDownload(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Book ID</font></th>'
        output+='<td><font size="" color="white">Rating</td>'
        rank = []
        rows = ""
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select book_name, avg(ratings) from feedback group by book_name")
            rows = cur.fetchall()
            
        
        wb = Workbook()
        ws = wb.active

    # Adding column headers
        headers = ['book_name', 'Average Ratings']
        ws.append(headers)

    # Adding data rows
        for row in rows:
            ws.append(row)

    # Create the response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Avgratings.xlsx'

    # Save the workbook to the response
        wb.save(response)
        
        return response
    
def CopiesSold(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Book Title</font></th>'
        output+='<td><font size="" color="white">Copies Sold</td>'
        rank = []
        rows = ""
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select b.book_name as Book_Title,res.occurrences as Copies_Sold from (SELECT product_id, COUNT(*) AS occurrences FROM ( SELECT TRIM(SUBSTRING_INDEX(SUBSTRING_INDEX(product_id, ',', n.digit + 1), ',', -1)) AS product_id FROM orders_new JOIN ( SELECT 0 AS digit UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3 UNION ALL SELECT 4 ) AS n ON CHAR_LENGTH(product_id) - CHAR_LENGTH(REPLACE(product_id, ',', '')) >= n.digit) AS extracted_product_ids GROUP BY product_id) as res left join addbook b on b.book_id=res.product_id")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                '</td></tr>'
        output += "</table><br/><br/><br/>"

        context= {'data': output}
        return render(request, 'CopiesSoldReport.html', context)  
    
def CopiesSoldDownload(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Book ID</font></th>'
        output+='<td><font size="" color="white">Rating</td>'
        rank = []
        rows = ""
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select b.book_name as Book_Title,res.occurrences as Copies_Sold from (SELECT product_id, COUNT(*) AS occurrences FROM ( SELECT TRIM(SUBSTRING_INDEX(SUBSTRING_INDEX(product_id, ',', n.digit + 1), ',', -1)) AS product_id FROM orders_new JOIN ( SELECT 0 AS digit UNION ALL SELECT 1 UNION ALL SELECT 2 UNION ALL SELECT 3 UNION ALL SELECT 4 ) AS n ON CHAR_LENGTH(product_id) - CHAR_LENGTH(REPLACE(product_id, ',', '')) >= n.digit) AS extracted_product_ids GROUP BY product_id) as res left join addbook b on b.book_id=res.product_id")
            rows = cur.fetchall()
            
        
        wb = Workbook()
        ws = wb.active

    # Adding column headers
        headers = ['book_name', 'Copies_Sold']
        ws.append(headers)

    # Adding data rows
        for row in rows:
            ws.append(row)

    # Create the response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=copies_sold.xlsx'

    # Save the workbook to the response
        wb.save(response)
        
        return response
    
def Topsales(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Order ID</font></th>'
        output+='<td><font size="" color="white">Book</td>'
        output+='<td><font size="" color="white">Amount</td>'
        output+='<td><font size="" color="white">OrderDate</td>'
        rank = []
        rows = ""
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("SELECT o.order_id,b.book_name,o.total_amount,o.order_date FROM orders_new o join addbook b on b.book_id = o.product_id WHERE order_date BETWEEN CURDATE() - INTERVAL 7 DAY AND CURDATE() ORDER BY total_amount DESC LIMIT 5")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
             
                output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                '</td></tr>'
        output += "</table><br/><br/><br/>"

        context= {'data': output}
        return render(request, 'TopSalesReport.html', context)  
    
def TopSalesDownload(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Book ID</font></th>'
        output+='<td><font size="" color="white">Rating</td>'
        rank = []
        rows = ""
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("SELECT o.order_id,b.book_name,o.username,o.total_amount,o.order_date FROM orders_new o join addbook b on b.book_id = o.product_id WHERE order_date BETWEEN CURDATE() - INTERVAL 7 DAY AND CURDATE() ORDER BY total_amount DESC LIMIT 5")
            rows = cur.fetchall()
            
        
        wb = Workbook()
        ws = wb.active

    # Adding column headers
        headers = ['Order_id','Book_Name','User','Amount','OrderDate']
        ws.append(headers)

    # Adding data rows
        for row in rows:
            ws.append(row)

    # Create the response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Topsales.xlsx'

    # Save the workbook to the response
        wb.save(response)
        
        return response
     
def ViewBooks(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Book ID</font></th>'
        output+='<td><font size="" color="white">Book Name</td>'
        output+='<td><font size="" color="white">Category</td>'
        output+='<td><font size="" color="white">Author Name</td>'
        output+='<td><font size="" color="white">Subject</td>'
        output+='<td><font size="" color="white">Cost</td>'
        output+='<td><font size="" color="white">Image</td></tr>'
        rank = []
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * FROM addbook")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                output+='<td><font size="" color="white">'+str(row[4])+'</td>'
                output+='<td><font size="" color="white">'+str(row[5])+'</td>'
                output+='<td><img src="static/books/'+row[6]+'" width="100" height="100" alt="" /></td></tr>'
        output += "</table><br/><br/><br/>"
        context= {'data': output}
        return render(request, 'AdminScreen.html', context)

def ViewRatings(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Username</font></th>'
        output+='<td><font size="" color="white">Book ID</td>'
        output+='<td><font size="" color="white">Ratings</td>'
        output+='<td><font size="" color="white">Feedback</td>'
        output+='<td><font size="" color="white">Date</td></tr>'
        rank = []
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * from feedback")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                output+='<td><font size="" color="white">'+str(row[4])+'</td></tr>'
        output += "</table><br/><br/><br/>"
        context= {'data': output}
        return render(request, 'AdminScreen.html', context)


def ManageOrders(request):
    if request.method == 'GET':
        output = '<table border=1><tr><th><font size="" color=white>Order ID</font></th>'
        output+='<td><font size="" color="white">Username</td>'
        output+='<td><font size="" color="white">Ordered Date</td>'
        output+='<td><font size="" color="white">Purchased Book ID</td>'
        output+='<td><font size="" color="white">Total Amount</td>'
        output+='<td><font size="" color="white">Order Status</td>'
        output+='<td><font size="" color="white">Complete Order</td></tr>'
        rank = []
        con = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        with con:
            cur = con.cursor()
            cur.execute("select * from orders_new where order_status='Pending'")
            rows = cur.fetchall()
            for row in rows:
                output+='<tr><td><font size="" color="white">'+str(row[0])+'</td>'
                output+='<td><font size="" color="white">'+str(row[1])+'</td>'
                output+='<td><font size="" color="white">'+str(row[2])+'</td>'
                output+='<td><font size="" color="white">'+str(row[3])+'</td>'
                output+='<td><font size="" color="white">'+str(row[4])+'</td>'
                output+='<td><font size="" color="white">'+str(row[5])+'</td>'
                output+='<td><a href=\'CompleteOrder?t1='+str(row[0])+'\'><font size=3 color=white>Click Here to Complete Order</font></a></td></tr>'
        output += "</table><br/><br/><br/>"
        context= {'data': output}
        return render(request, 'AdminScreen.html', context)

def CompleteOrder(request):
    if request.method == 'GET':
        order_id = request.GET.get('t1', False)
        db_connection = pymysql.connect(host='127.0.0.1',port = 3306,user = 'root', password = 'root', database = 'onlinebookstore',charset='utf8')
        db_cursor = db_connection.cursor()
        student_sql_query = "update orders_new set order_status='Completed' where order_id='"+order_id+"'"
        db_cursor.execute(student_sql_query)
        db_connection.commit()
        status = 'Ordered Completed Successfully'
        context= {'data':status}
        return render(request, 'AdminScreen.html', context)
        
